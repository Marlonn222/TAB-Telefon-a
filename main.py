# Codigo Principal Del Boot De AproTelefonica
# Improto las librerias requeridas
import pyautogui  # Importa el módulo pyautogui para realizar automatización de la interfaz gráfica del usuario (GUI).
import shutil # Importa el módulo shutil para operaciones de manipulación de archivos y directorios.
import asyncio # Importa el módulo asyncio para admitir el uso de corutinas y tareas asincrónicas.
import sys # Importa el módulo sys para acceder a la funcionalidad del intérprete de Python.
import os # Importa el módulo os para interactuar con el sistema operativo.
import time # Importa el modulo time que proporciona funciones relacionadas con el tiempo y el reloj.
import openpyxl # Importa el módulo openpyxl para manejar datos de entrada a través de Excel.
import logging
# Importor programas creados con sud procesos
from Opensistem import (openCRM,connectToCRM)
from searchandupdateot import (searchAproTelefono)
from closeapps import (closeCRM)
from storagefunctions import (showDesktop,make_noise,get_username_os,terminateProcess)
from telegramfunctions import (sendTelegramMsg)
from datetime import datetime
from time import sleep
from dotenv import load_dotenv

async def main():   
    
    #while True:
    # print(pyautogui.position())    

    load_dotenv() 

    # CONSTANTS
    INPUT_DIRECTORY = 'C:\AproTelefonica\Insumos'
    INPUT_FILENAME = 'Input BOT 001 V1.xlsx'
    SUPER_LOG_FILENAME = 'C:/AproTelefonica/logs/super_log.txt'
    GENERIC_ERROR_MSG = 'No Se ha procesado el incidente - Para ver mas detalles ver el archivo de logs'
    # VARIABLES        
    counter = 1
    total_rows = 0
    total_cols = 0
    resp_funcion = None
    
    try:
        # Inicio o Home del sistema
        FORMAT ='%(asctime)s - %(name)s - %(user)s - %(levelname)s - %(message)s'
        logging.basicConfig(filename=SUPER_LOG_FILENAME, filemode='w',format=FORMAT, datefmt='%m/%d/%Y %I:%M:%S %p',level=logging.WARNING)
        attrs = {'user': get_username_os()}
        
        # Lobbi del sistema antes de comenzar
        print("----------------- BOT PROCESO DE APROVICIONAMIENTO TELEFONICO V1 --------------------")                        
        print("Ingresa el valor de 1 para inciar y cualquier otro valor para salir")
        actividad_rpa = input("1. inciar proceso aprovicionamiento telefonia\n2. No se realizara actividad\n")
        if actividad_rpa == "1":
            actividad_rpa_selected = 'Procese de aprovicionamiento Telefonia'
        else:
            print("No se realizara actividad gracias por avisar")            
            print("Exit")
            exit_program()
        
        # Proceso de Estracion de Datos de la Base (Excel), Extraer datos de un archivo Excel y abrir un archivo Excel existente       
        file = os.path.join(INPUT_DIRECTORY,INPUT_FILENAME)
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        # Eliminar la hoja 'Input Backup' si existe
        if 'Input Backup' in wb.sheetnames:
            del wb['Input Backup']
        
        # Crear una copia de seguridad de la hoja de trabajo de Excel y notificación de inicio de ejecución
        target_ws = wb.copy_worksheet(ws)
        target_ws.title = "Input Backup"
        total_rows = len(ws['A'])
        total_cols = len(ws[1])
        print("Total de registros en base fuente: ",total_rows-1)
        print("Total de columnas en base fuente: ",total_cols)
        
        # Hace el proceso de apertura y coneccion al sistema CRM        
        openCRM()
        sleep(1)
        connectToCRM()       
        
        # Procesamiento de datos en el bucle
        rows = ws.iter_rows(min_row=2, max_row=total_rows, min_col=1, max_col=total_cols)    
       
        for row in rows:
            # se omiten aquellos registros que en la col STATUS tenga CORRECTO o contenga la palabra ERROR
            if row[1].value == 'CORRECTO' or str(row[1].value).find("ERROR") != -1:
                continue

            # se formatean las casillas seleccionados y se hace el llamado a la funcion corespondiente que ejcuta el proceso
            if actividad_rpa.lower() == "1":   
                print('Validando registro: ',str(row[0].value))                
                resp_funcion = searchAproTelefono(str(row[0].value))
                print("Respuesta de la función:",resp_funcion)     

            # Manejo de las posibles respuestas identificadas en los metodos de control e impresion en estatus excel.
            if resp_funcion == 0:
                ot_completadas = [str(row[0].value)]
                row[1].value = 'CORRECTO'
            elif resp_funcion == 14:
                row[1].value = 'ERROR - No Cuenta con OTH_Configuracion'
            elif resp_funcion == 13:
                row[1].value = 'ERROR - Numero de telefono no presente'
            elif resp_funcion == 12:
                row[1].value = 'ERROR - Numero de telefono no concide'
            elif resp_funcion == 11:
                row[1].value = 'ERROR - OT CANCELADA'
            elif resp_funcion == 10:
                row[1].value = 'ERROR - No se identifica vista de detalles de OT reconocida en el CRM'                
            elif resp_funcion == 9:             
                row[1].value = 'ERROR - Se presenta mensaje de advertencia en el proceso en CRM'                    
            else:              
                row[1].value = 'ERROR - '+ GENERIC_ERROR_MSG              
                terminateProcess('CRM.exe')
                print("Se ha terminado el proceso CRM.exe")
                sleep(2)
                openCRM()            
                connectToCRM()
                
            print("Registro(counter) #: ",counter)
            if counter % 10 == 0:
                await sendTelegramMsg('Se han procesado '+ str(counter) +' registros y se procede a reiniciar CRM ')
                terminateProcess('CRM.exe')
                print("Se ha terminado el proceso CRM.exe")
                sleep(2)
                openCRM()
                connectToCRM()
                                            
            counter = counter + 1
            wb.save(file) 
                       
        print("Fin del Proceso Macro")

    except FileNotFoundError as e:
        
        exit_program()
        
    except Exception as e:
        
        print(f"An error occurred: {e}")     
        e_type, e_object, e_traceback = sys.exc_info()
        e_line_number = e_traceback.tb_lineno
        print (sys.exc_info())
        exit_program()
    
    else:
        # Hace el proceso de cierre y finalizacion de la coneccion al sistema CRM   
        closeCRM()
        sleep(1)
        showDesktop()                                        
        exit_program()        
    finally: 
        sleep(1)
        make_noise() 
        
def exit_program():
    sys.exit(0)    
    
if __name__ == "__main__":
   asyncio.run(main())